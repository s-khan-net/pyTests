import openpyxl

wb = openpyxl.load_workbook("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]
# cell = sheet["a1"]
cell = sheet.cell(row=1, column=1)
print(cell.value)

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column):
        cell = sheet.cell(row, column)
        print(cell.value)
#col wise9
cells = sheet["a:c"]
print("all cells in cols a to c as tuple of tuples - >",cells)
#row wise
cells = sheet[1:3]
print("all rows from 1 to 3 as tuple of tuples->",cells)